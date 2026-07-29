#!/usr/bin/env python3
"""Convierte un backlog Excel existente en JSON importable por Project Mapper.

Lee un .xlsx con columnas mínimas (nombre, tipo, sección, owner, deadline) y
emite un JSON con la forma del Store de Project Mapper. Las secciones B-F del
cuestionario (Impacto, Esfuerzo, Urgencia/Importancia, información) no vienen
en un backlog típico, así que se completan con valores neutros y F1=1
(confianza mínima) + F2 marcando esos campos como conjetura, de forma que la
actividad importada aparezca con el badge "necesita info" en el Eisenhower
hasta que alguien complete el cuestionario de verdad.

Uso:
    python import_backlog.py backlog.xlsx -o import.json
    python import_backlog.py backlog.xlsx --sheet "Hoja1"

El JSON de salida se puede usar de dos formas (§2.1 de la propuesta: el
script es opcional, la app funciona sin él):
    1. Subirlo en Configuración > Importar backlog (o `POST /import` con ese
       archivo en el campo "file") — agrega las actividades a un store
       existente sin tocar secciones/pesos/umbrales ya configurados.
    2. Reemplazar directamente el archivo de datos (`-data`) de la app con
       este JSON — solo recomendado si todavía no hay datos reales, porque
       reemplaza también secciones/pesos/umbrales.
"""

from __future__ import annotations

import argparse
import datetime
import json
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("Falta openpyxl. Instálalo con: pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)

DEFAULT_SECTIONS = ["Operaciones", "Finanzas", "Legal", "IT"]
DEFAULT_WEIGHTS = {
    "impact": {"B1": 0.4, "B2": 0.2, "B3": 0.25, "B4": 0.15},
    "effort": {"C1": 0.4, "C2": 0.2, "C3": 0.25, "C4": 0.15},
}
DEFAULT_THRESHOLDS = {"impact": 50, "effort": 50, "urgency": 50, "importance": 50}

# Sinónimos de encabezado aceptados (todo se compara en minúsculas, sin acentos).
HEADER_SYNONYMS = {
    "name": ["nombre", "name", "actividad", "proyecto"],
    "type": ["tipo", "type"],
    "section": ["seccion", "sección", "section", "area", "área"],
    "owner": ["owner", "responsable", "dueño", "dueno"],
    "deadline": ["deadline", "fecha limite", "fecha límite", "fecha_limite", "due", "vencimiento"],
}

TYPE_ALIASES = {
    "project": "Project", "proyecto": "Project",
    "workstream": "Workstream", "iniciativa": "Workstream",
    "recurring": "Recurring", "recurrente": "Recurring",
    "adhoc": "AdHoc", "ad-hoc": "AdHoc", "ad hoc": "AdHoc", "tarea": "AdHoc", "tarea puntual": "AdHoc",
}

# Respuestas neutras para las secciones B-F que un backlog no trae. F1=1 y F2
# marcando estos mismos campos como conjetura es intencional: es la única
# forma honesta de reflejar que el script, no la persona dueña de la
# actividad, eligió estos valores — activa el flag "necesita info".
GUESSED_KEYS = ["B1", "B2", "B3", "B4", "C1", "C2", "C3", "C4", "D2", "D3", "D4"]


def _norm(s: str) -> str:
    return s.strip().lower()


def find_columns(header_row: list) -> dict:
    """Mapea nombre lógico ('name', 'type', ...) -> índice de columna (0-based)."""
    normalized = [_norm(str(c)) if c is not None else "" for c in header_row]
    columns = {}
    for logical, synonyms in HEADER_SYNONYMS.items():
        for idx, cell in enumerate(normalized):
            if cell in synonyms:
                columns[logical] = idx
                break
    return columns


def normalize_type(raw: str | None, row_num: int, warnings: list) -> str:
    if not raw:
        return "AdHoc"
    key = _norm(str(raw))
    if key in TYPE_ALIASES:
        return TYPE_ALIASES[key]
    warnings.append(f"Fila {row_num}: tipo {raw!r} no reconocido, se usa 'AdHoc'.")
    return "AdHoc"


def normalize_deadline(raw, row_num: int, warnings: list) -> str | None:
    if raw is None or raw == "":
        return None
    if isinstance(raw, (datetime.date, datetime.datetime)):
        return raw.strftime("%Y-%m-%d")
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(str(raw), fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    warnings.append(f"Fila {row_num}: no se pudo interpretar la fecha {raw!r}, se importa sin deadline.")
    return None


def build_activity(name: str, atype: str, section: str, owner: str, deadline: str | None) -> dict:
    return {
        "name": name,
        "type": atype,
        "owner": owner or "",
        "section": section or "",
        "involved": [],
        "description": "",
        "answers": {
            "b1": 3, "b2": 3, "b3": 3, "b4": 3,
            "c1": "20-60", "c2": 1, "c3": 3, "c4": "none",
            "d1": deadline, "d2": "nothing", "d3": "mine", "d4": "yes",
            "f1": 1, "f2": GUESSED_KEYS,
            "f3": "Importado automáticamente desde backlog Excel; faltan las respuestas detalladas de Impacto/Esfuerzo/Urgencia. Completar el cuestionario para reemplazar estos valores.",
        },
        "deliverables": [],
    }


def convert(xlsx_path: str, sheet_name: str | None) -> tuple[dict, list]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("la hoja está vacía")

    columns = find_columns(rows[0])
    if "name" not in columns:
        raise ValueError(
            "no se encontró una columna de nombre (se aceptan: "
            + ", ".join(HEADER_SYNONYMS["name"]) + ")"
        )

    warnings: list[str] = []
    activities = []
    for row_num, row in enumerate(rows[1:], start=2):
        def get(logical):
            idx = columns.get(logical)
            return row[idx] if idx is not None and idx < len(row) else None

        name = get("name")
        if not name or not str(name).strip():
            continue  # fila vacía o sin nombre: se ignora, no es un error
        atype = normalize_type(get("type"), row_num, warnings)
        deadline = normalize_deadline(get("deadline"), row_num, warnings)
        activities.append(build_activity(str(name).strip(), atype, get("section") or "", get("owner") or "", deadline))

    store = {
        "sections": DEFAULT_SECTIONS,
        "weights": DEFAULT_WEIGHTS,
        "thresholds": DEFAULT_THRESHOLDS,
        "activities": activities,
    }
    return store, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("xlsx", help="ruta al backlog .xlsx de entrada")
    parser.add_argument("-o", "--output", help="ruta del JSON de salida (default: <nombre>.json)")
    parser.add_argument("--sheet", help="nombre de la hoja a leer (default: la primera)")
    args = parser.parse_args()

    output = args.output or (args.xlsx.rsplit(".", 1)[0] + ".json")

    try:
        store, warnings = convert(args.xlsx, args.sheet)
    except Exception as exc:  # noqa: BLE001 - script de línea de comandos
        print(f"Error: {exc}", file=sys.stderr)
        return 1

    with open(output, "w", encoding="utf-8") as f:
        json.dump(store, f, indent=2, ensure_ascii=False)

    print(f"{len(store['activities'])} actividad(es) escritas en {output}")
    for w in warnings:
        print(f"  aviso: {w}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
